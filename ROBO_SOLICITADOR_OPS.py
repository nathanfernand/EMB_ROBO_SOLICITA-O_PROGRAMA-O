# =============================================================================
# ROBÔ SAP — ZLOLMM025
# =============================================================================
# Descrição : Automação de solicitação de ordens de produção no SAP GUI.
#             Para cada linha da planilha DataBase, o robô identifica o tipo
#             de carro (Montador ou Fabricante) e executa a transação adequada.
#
# Fluxo de execução:
#   1. Usuário clica em "Login SAP" na interface gráfica (Tkinter)
#   2. O robô abre o SAP e realiza o login
#   3. Para cada linha da planilha:
#      ├─ Carro em TIPO_CARRO → ZDPQPL126_MONTADOR  (linha M)
#      └─ Carro fora do dict  → ZDPQPL126_FABRICANTE (linha F)
#           └─ Falha?         → ZLOLMM027_MTS        (fallback)
#   4. Ao final, exibe o resumo na janela e abre o Excel atualizado
#
# Arquivo de entrada : C:\base\ZLOLMM025.xlsx  (aba "DataBase")
# Arquivo intermediário: C:\base\ZDPQPL126.txt (exportado pelo SAP)
# =============================================================================

# --- Bibliotecas padrão -------------------------------------------------------
import csv        # Leitura do arquivo TXT exportado pelo SAP
import os         # Verificação de existência de arquivos e abertura do Excel
import subprocess # Abertura do executável do SAP
import sys        # Captura de informações de exceções no login
import time       # Delays necessários para sincronização com o SAP

# --- Interface gráfica --------------------------------------------------------
from tkinter import Button, Label, Tk, mainloop, messagebox

# --- Bibliotecas de terceiros -------------------------------------------------
from openpyxl import load_workbook
import pandas as pd          # Leitura e escrita da planilha Excel
import win32com.client       # Comunicação COM com o SAP GUI Scripting API

# =============================================================================
# DICIONÁRIO DE TIPOS DE CARRO
# =============================================================================
# Mapeia o código do carro para seu tipo de processo.
#
# Regra de roteamento:
#   • Código PRESENTE neste dicionário → processa via ZDPQPL126_MONTADOR
#   • Código AUSENTE  neste dicionário → processa via ZDPQPL126_FABRICANTE
#
# Para adicionar novos tipos Montador, basta inserir a chave neste dicionário.
# =============================================================================
TIPO_CARRO = {
    "I1K": "Montador",
    "I1M": "Montador",
    "I1O": "Montador",
    "I1Q": "Montador",
    "I1S": "Montador",
    "I2K": "Montador",
    "I2O": "Montador",
    "I2Q": "Montador",
    "I2S": "Montador",
    "I2Y": "Montador",
    "I3J": "Montador",
    "I3Q": "Montador",
    "I3S": "Montador",
    "I4S": "Montador",
    "I4Y": "Montador",
    "IA2": "Montador",
    "IAR": "Montador",
    "IAZ": "Montador",
    "IB2": "Montador",
    "IBJ": "Montador",
    "I4J": "Montador",
    "I5Q": "Montador",
    "IA3": "Montador",
    "IA4": "Montador",
    "IAK": "Montador",
    "IAL": "Montador",
    "IAM": "Montador",
    "IAN": "Montador",
    "IAP": "Montador",
    "IAT": "Montador",
    "IAU": "Montador",
    "IAV": "Montador",
    "IAW": "Montador",
    "ICJ": "Montador",
    "IDC": "Montador",
    "IL1": "Montador",
    "IL2": "Montador",
    "IT1": "Montador",
    "IT2": "Montador",
    "IAX": "Montador",
    "IBK": "Montador",
    "IW1": "Montador",
    "I2G": "Montador",
    "I1V": "Montador",
    "I2J": "Montador",
    "I1J": "Montador",
    "I3K": "Montador",
    "I4K": "Montador",
    "I1Y": "Montador",
    "I3Y": "Montador",
}

# =============================================================================
# DICIONÁRIO DE CARROS — ZLOLMM027
# =============================================================================
# Mapeia os códigos de carro que devem ser processados via ZLOLMM027.
# =============================================================================
CARRO_ZLOLMM027 = {
    "MMU": "ZLOLMM027",
    "I1D": "ZLOLMM027",
    "MMX": "ZLOLMM027",
    "MMW": "ZLOLMM027",
    "MMY": "ZLOLMM027",
    "MMZ": "ZLOLMM027",
    "Z1R": "ZLOLMM027",
    "ZAR": "ZLOLMM027",
    "KVC": "ZLOLMM027",
    "KVA": "ZLOLMM027",
    "JAK": "ZLOLMM027",
    "JAD": "ZLOLMM027",
    "JAB": "ZLOLMM027",
    "JAJ": "ZLOLMM027",
    "JAI": "ZLOLMM027",
    "JAG": "ZLOLMM027",
    "JAH": "ZLOLMM027",
    "JAF": "ZLOLMM027",
    "JAE": "ZLOLMM027",
    "JAC": "ZLOLMM027",
    "SO9": "ZLOLMM027",
    "SO6": "ZLOLMM027",
    "SO5": "ZLOLMM027",
    "SO4": "ZLOLMM027",
    "SO3": "ZLOLMM027",
    "SO2": "ZLOLMM027",
    "SO1": "ZLOLMM027",
    "KS1": "ZLOLMM027",
    "KS2": "ZLOLMM027",
    "KS4": "ZLOLMM027",
    "KS5": "ZLOLMM027",
    "KS3": "ZLOLMM027",
    "KS6": "ZLOLMM027",
    "KS7": "ZLOLMM027",
    "KS8": "ZLOLMM027",
    "KC2": "ZLOLMM027",
}


# =============================================================================
# CLASSE PRINCIPAL — SapGui
# =============================================================================
class SapGui(object):
    """
    Classe responsável por toda a automação do SAP GUI.

    Responsabilidades:
        - Conectar ao SAP via interface COM (win32com)
        - Ler e gravar a planilha de controle (ZLOLMM025.xlsx)
        - Rotear cada ordem para a transação correta conforme o tipo de carro
        - Atualizar a coluna STATUS de cada linha ao finalizar o processamento

    Atributos de classe:
        SAP_GUI_PATH    (str): Caminho do executável do SAP Logon Pad.
        EXCEL_PATH      (str): Caminho da planilha de controle.
        TXT_EXPORT_PATH (str): Caminho do arquivo TXT exportado pelo SAP.
    """

    # -------------------------------------------------------------------------
    # CAMINHOS DOS RECURSOS
    # -------------------------------------------------------------------------
    SAP_GUI_PATH    = r"C:\Program Files (x86)\SAP\FrontEnd\SAPgui\saplgpad.exe"
    EXCEL_PATH      = r"C:\base\ZLOLMM025.xlsx"
    TXT_EXPORT_PATH = r"C:\base\ZDPQPL126.txt"

    # =========================================================================
    # INICIALIZAÇÃO E CONEXÃO COM O SAP
    # =========================================================================
    def __init__(self):
        """
        Conecta ao SAP GUI Scripting

        Tenta primeiro reutilizar uma sessão SAP já aberta e autenticada.
        Se o SAP não estiver aberto (ou nenhuma conexão ativa for encontrada),
        abre o SAP Logon Pad e cria uma nova conexão normalmente.

        Fluxo:
            1. [REUTILIZAÇÃO] Tenta obter o objeto COM "SAPGUI" já em memória.
               Se houver ao menos uma conexão e uma sessão ativas, usa-as
               diretamente sem abrir uma nova janela.
            2. [ABERTURA NORMAL] Se o passo 1 falhar (SAP fechado, sem conexão
               ou sem sessão), abre o executável do SAP, aguarda carregar,
               e cria uma nova conexão com o servidor.

        Em caso de falha total, exibe uma caixa de erro e registra no console.
        """
        self.path = self.SAP_GUI_PATH
        # Buffer de status que falharam ao persistir no Excel para nova tentativa no fim.
        self._status_pendentes = {}

        # ------------------------------------------------------------------
        # PASSO 1 — Tenta reutilizar sessão SAP já aberta
        # ------------------------------------------------------------------
        try:
            self.SapGuiAuto = win32com.client.GetObject("SAPGUI")
            application = self.SapGuiAuto.GetScriptingEngine

            # Verifica se existe pelo menos uma conexão com pelo menos uma sessão
            if application.Children.Count > 0:
                self.connection = application.Children(0)
                if self.connection.Children.Count > 0:
                    self.session = self.connection.Children(0)
                    self.session.findById("wnd[0]").maximize()
                    print("✓ Sessão SAP existente reutilizada — novo login não necessário.")
                    return  # Sai do __init__; sapLogin() tratará os popups residuais

        except Exception:
            # SAP não está aberto ou não é acessível via COM — segue para abertura normal
            pass

        # ------------------------------------------------------------------
        # PASSO 2 — SAP não estava disponível: abre e conecta normalmente
        # ------------------------------------------------------------------
        try:
            # Abre o executável do SAP em segundo plano
            subprocess.Popen(self.path)
            time.sleep(1)  # Aguarda o SAP Logon Pad carregar completamente

            # Obtém o objeto raiz do SAP GUI Scripting (interface COM)
            self.SapGuiAuto = win32com.client.GetObject("SAPGUI")
            application = self.SapGuiAuto.GetScriptingEngine

            # Abre a conexão com o servidor SAP especificado no Logon Pad
            self.connection = application.OpenConnection("02- EBP - SAP Corp (FI/CO)", True)
            time.sleep(1.5)  # Aguarda a conexão ser estabelecida

            # Obtém a primeira sessão aberta e maximiza a janela principal
            self.session = self.connection.Children(0)
            self.session.findById("wnd[0]").maximize()
            print("✓ Nova sessão SAP aberta e conectada.")

        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao conectar ao SAP: {e}")
            print(f"Erro COM: {e}")

    # =========================================================================
    # MÉTODOS AUXILIARES (HELPERS)
    # =========================================================================

    def _reset_transacao(self):
        """
        Retorna à tela inicial do SAP enviando o comando /N.

        Deve ser chamado após cada ordem processada (com sucesso ou falha)
        para garantir que o SAP esteja em estado limpo antes da próxima iteração.
        """
        self.session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
        self.session.findById("wnd[0]").sendVKey(0)

    def _obter_coluna_status(self, worksheet):
        """
        Localiza a coluna STATUS na planilha, criando o cabeçalho se necessário.

        Retorna:
            int: Índice 1-based da coluna STATUS.
        """
        for coluna in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=1, column=coluna).value == "STATUS":
                return coluna

        coluna_status = worksheet.max_column + 1
        worksheet.cell(row=1, column=coluna_status).value = "STATUS"
        return coluna_status

    def _salvar_planilha_via_excel_com(self, data, arquivo_excel):
        """
        Persiste a coluna STATUS via automação COM do Excel.

        Este caminho é usado como fallback quando o arquivo está aberto e
        bloqueado para escrita por openpyxl.

        Retorna:
            bool: True se salvou com sucesso, False em caso de erro.
        """
        excel = None
        workbook = None
        criou_instancia_excel = False
        workbook_ja_aberto = False

        try:
            try:
                excel = win32com.client.GetObject(Class="Excel.Application")
            except Exception:
                excel = win32com.client.Dispatch("Excel.Application")
                criou_instancia_excel = True

            caminho_alvo = os.path.abspath(arquivo_excel).lower()
            for wb in excel.Workbooks:
                if os.path.abspath(wb.FullName).lower() == caminho_alvo:
                    workbook = wb
                    workbook_ja_aberto = True
                    break

            if workbook is None:
                workbook = excel.Workbooks.Open(arquivo_excel)

            worksheet = workbook.Worksheets("DataBase")

            coluna_status = None
            total_colunas = max(worksheet.UsedRange.Columns.Count, 1)
            for coluna in range(1, total_colunas + 1):
                if worksheet.Cells(1, coluna).Value == "STATUS":
                    coluna_status = coluna
                    break

            if coluna_status is None:
                coluna_status = total_colunas + 1
                worksheet.Cells(1, coluna_status).Value = "STATUS"

            for index, status_msg in data["STATUS"].items():
                worksheet.Cells(index + 2, coluna_status).Value = status_msg

            workbook.Save()
            print("✓ STATUS salvo via Excel COM.")
            return True

        except Exception as excel_com_error:
            print(f"Falha ao atualizar STATUS no Excel via COM: {excel_com_error}")
            return False
        finally:
            if workbook is not None and not workbook_ja_aberto:
                workbook.Close(SaveChanges=True)
            if excel is not None and criou_instancia_excel:
                excel.Quit()

    def _salvar_planilha(self, data, arquivo_excel):
        """
        Persiste somente a coluna STATUS do DataFrame na aba 'DataBase' do Excel.

        Parâmetros:
            data (DataFrame): DataFrame com os dados já atualizados em memória.
            arquivo_excel (str): Caminho completo do arquivo Excel de destino.

        Retorna:
            bool: True se o arquivo foi salvo com sucesso, False em caso de erro.
        """
        for tentativa in range(1, 6):
            workbook = None
            try:
                workbook = load_workbook(arquivo_excel)
                worksheet = workbook["DataBase"]
                coluna_status = self._obter_coluna_status(worksheet)

                for index, status_msg in data["STATUS"].items():
                    worksheet.cell(row=index + 2, column=coluna_status).value = status_msg

                workbook.save(arquivo_excel)
                return True
            except PermissionError as excel_error:
                print(
                    f"Falha ao atualizar STATUS no Excel (tentativa {tentativa}/5): "
                    f"arquivo em uso/bloqueado. Feche o Excel e tente novamente. Erro: {excel_error}"
                )
                if self._salvar_planilha_via_excel_com(data, arquivo_excel):
                    return True
                time.sleep(2)
            except Exception as excel_error:
                print(f"Falha ao atualizar STATUS no Excel (tentativa {tentativa}/5): {excel_error}")
                time.sleep(2)
            finally:
                if workbook is not None:
                    workbook.close()

        return False

    def _ler_status_excel(self, arquivo_excel, index):
        """
        Lê o valor atual da célula STATUS diretamente do Excel para diagnóstico.

        Retorna:
            str | None: Valor encontrado na célula STATUS da linha informada.
        """
        workbook = None
        try:
            workbook = load_workbook(arquivo_excel, data_only=True)
            worksheet = workbook["DataBase"]
            coluna_status = self._obter_coluna_status(worksheet)
            return worksheet.cell(row=index + 2, column=coluna_status).value
        except Exception as readback_error:
            print(f"⚠ Falha ao reler STATUS no Excel para diagnóstico: {readback_error}")
            return None
        finally:
            if workbook is not None:
                workbook.close()

    def _atualizar_status(self, data, index, arquivo_excel, status_msg):
        """
        Atualiza a coluna STATUS da linha especificada e salva no Excel.

        Parâmetros:
            data (DataFrame)  : DataFrame em memória.
            index (int)       : Índice da linha a ser atualizada.
            arquivo_excel (str): Caminho do arquivo Excel.
            status_msg (str)  : Texto a ser gravado na coluna STATUS.

        Retorna:
            bool: Resultado de _salvar_planilha().
        """
        carro = data.at[index, "CARRO"] if "CARRO" in data.columns else "<sem CARRO>"
        op = data.at[index, "OP"] if "OP" in data.columns else "<sem OP>"
        status_anterior = data.at[index, "STATUS"] if "STATUS" in data.columns else None

        print(
            f"[STATUS] Linha={index + 1} CARRO={carro} OP={op} | "
            f"anterior={status_anterior!r} -> novo={status_msg!r}"
        )

        data.at[index, "STATUS"] = status_msg
        status_salvo = self._salvar_planilha(data, arquivo_excel)

        if not status_salvo:
            # Mantem o último status por linha para tentativa posterior de persistência.
            self._status_pendentes[index] = status_msg
        else:
            self._status_pendentes.pop(index, None)

        status_memoria = data.at[index, "STATUS"] if "STATUS" in data.columns else None
        status_excel = self._ler_status_excel(arquivo_excel, index)

        print(
            f"[STATUS] Linha={index + 1} resultado salvar={status_salvo} | "
            f"memoria={status_memoria!r} | excel={status_excel!r} | "
            f"pendentes={self._status_pendentes}"
        )

        return status_salvo

    def _persistir_status_pendentes(self, data, arquivo_excel):
        """
        Reaplica e persiste no Excel os STATUS que falharam em tentativas anteriores.

        Retorna:
            bool: True quando todos os pendentes foram gravados, False caso permaneçam falhas.
        """
        if not self._status_pendentes:
            return True

        print(f"↻ Tentando persistir {len(self._status_pendentes)} STATUS pendente(s)...")
        for index, status_msg in self._status_pendentes.items():
            data.at[index, "STATUS"] = status_msg

        status_salvo = self._salvar_planilha(data, arquivo_excel)
        if status_salvo:
            self._status_pendentes.clear()
            print("✓ STATUS pendentes persistidos com sucesso.")
            return True

        print("⚠ Ainda há STATUS pendentes não gravados no Excel.")
        return False

    def _tratar_falha_e_continuar(self, data, index, arquivo_excel, data_126, status_msg, log_msg):
        """
        Tratamento padronizado de falhas durante o processamento de uma linha.

        Ações executadas (nesta ordem):
            1. Imprime a mensagem de log no console
            2. Grava o STATUS de erro na linha do Excel
            3. Limpa a lista temporária data_126
            4. Reseta a transação do SAP para o estado inicial (/N)

        Parâmetros:
            data (DataFrame)   : DataFrame em memória.
            index (int)        : Índice da linha com falha.
            arquivo_excel (str): Caminho do arquivo Excel.
            data_126 (list)    : Lista temporária de valores a ser limpa.
            status_msg (str)   : Mensagem de status a gravar no Excel.
            log_msg (str)      : Mensagem a imprimir no console.
        """
        print(log_msg)
        print(f"[FALHA] Linha={index + 1} | status_msg={status_msg!r}")
        status_salvo = self._atualizar_status(data, index, arquivo_excel, status_msg)
        if not status_salvo:
            print(f"⚠ Não foi possível persistir STATUS da linha {index + 1} no Excel.")
        data_126.clear()
        self._reset_transacao()

    def _rotear_para_zlolmm027(self, row, index, data, arquivo_excel):
        """
        Aciona ZLOLMM027_MTS e atualiza o STATUS em caso de sucesso.

        Centraliza o roteamento para ZLOLMM027 tanto no caso direto (Regra 1 —
        carro presente em CARRO_ZLOLMM027) quanto no re-roteamento pós-falha de
        ZDPQPL126_FABRICANTE (Regra 2). O STATUS de erro, se houver, é definido
        internamente por ZLOLMM027_MTS.

        Parâmetros:
            row           (Series)    : Linha atual do DataFrame.
            index         (int)       : Índice da linha no DataFrame.
            data          (DataFrame) : DataFrame completo.
            arquivo_excel (str)       : Caminho do arquivo Excel.
        """
        try:
            self.ZLOLMM027_MTS(row, index, data)
            status_msg = "ORDEM SOLICITADA VIA ZLOLMM027"
            self._atualizar_status(data, index, arquivo_excel, status_msg)
            print(f"✓ ZLOLMM027_MTS concluído com sucesso")
            print(f"✓ Status atualizado: {status_msg}")
        except Exception as e_mts:
            print(f"❌ Erro em ZLOLMM027_MTS: {e_mts}")
            # Status já definido internamente por ZLOLMM027_MTS

    # =========================================================================
    # ORQUESTRADOR PRINCIPAL
    # =========================================================================


    def processar_ordem_com_fluxo(self):
        """
        Orquestrador central do robô. Lê todas as linhas da planilha e roteia
        cada ordem para a função de processamento correta com base no tipo de carro.

        Fluxo de decisão por linha (Regra 1 + Regra 2):
        ┌──────────────────────────────────────────────────────────────────────┐
        │ CARRO em TIPO_CARRO?     → ZDPQPL126_MONTADOR                        │
        │ CARRO em CARRO_ZLOLMM027?→ ZLOLMM027_MTS  (roteamento direto)       │
        │ Nenhum dos anteriores   → ZDPQPL126_FABRICANTE                       │
        │   └─ Falha?  Re-verifica Regra 1:                                    │
        │       CARRO em CARRO_ZLOLMM027? → ZLOLMM027_MTS                      │
        │       Não?                      → status definido por FABRICANTE     │
        └──────────────────────────────────────────────────────────────────────┘

        Cada função chamada é responsável por definir sua própria mensagem de
        STATUS em caso de erro. Este orquestrador apenas roteia e consolida.

        Retorna:
            tuple(int, int, int): (total, sucesso, erros) contados ao final
                                  pela releitura do Excel já persistido.
        """
        arquivo_excel = self.EXCEL_PATH

        # Verifica se o arquivo de controle existe antes de qualquer operação
        if not os.path.exists(arquivo_excel):
            messagebox.showerror("Erro", f"Arquivo não encontrado: {arquivo_excel}")
            return

        try:
            # Lê o DataFrame completo da aba DataBase, convertendo tudo para str
            data = pd.read_excel(arquivo_excel, sheet_name="DataBase").astype(str)

            print("\n" + "=" * 80)
            print("INICIANDO PROCESSAMENTO DE ORDENS")
            print("=" * 80)

            # -----------------------------------------------------------------
            # LOOP PRINCIPAL — itera linha a linha; cada função processa apenas
            # a linha recebida e retorna. O próximo índice é processado aqui.
            # -----------------------------------------------------------------
            for index, row in data.iterrows():
                try:
                    carro = row["CARRO"].strip()
                    print(f"\n[Linha {index + 1}] Processando CARRO: {carro}")

                    # ----------------------------------------------------------
                    # ROTEAMENTO: categoriza o carro e aciona a função adequada
                    # ----------------------------------------------------------
                    if carro in TIPO_CARRO:
                        # ── MONTADOR ──────────────────────────────────────────
                        print(f"✓ '{carro}' encontrado em TIPO_CARRO → MONTADOR")
                        try:
                            self.ZDPQPL126_MONTADOR(row, index, data)
                            print(f"✓ ZDPQPL126_MONTADOR concluído com sucesso")
                        except Exception as e:
                            print(f"❌ Erro em ZDPQPL126_MONTADOR: {e}")
                            # Status já definido internamente por ZDPQPL126_MONTADOR

                    elif carro in CARRO_ZLOLMM027:
                        # ── ZLOLMM027_MTS — roteamento direto (Regra 1) ───────
                        print(f"✓ '{carro}' encontrado em CARRO_ZLOLMM027 → ZLOLMM027_MTS")
                        self._rotear_para_zlolmm027(row, index, data, arquivo_excel)

                    else:
                        # ── FABRICANTE ────────────────────────────────────────
                        print(f"✗ '{carro}' NÃO encontrado em TIPO_CARRO nem CARRO_ZLOLMM027 → FABRICANTE")
                        try:
                            sucesso_fabricante = self.ZDPQPL126_FABRICANTE(row, index, data)
                            if sucesso_fabricante:
                                print(f"✓ ZDPQPL126_FABRICANTE concluído com sucesso")
                            else:
                                print("❌ ZDPQPL126_FABRICANTE finalizado com falha (sem fallback para ZLOLMM027_MTS)")
                        except Exception as e_fabricante_nao_mapeado:
                            status_msg = "ORDEM NÃO SINCRONIZADA NO MES, ACIONAR PPCP"
                            print(f"❌ Exceção não mapeada no FABRICANTE: {e_fabricante_nao_mapeado}")
                            self._atualizar_status(data, index, arquivo_excel, status_msg)
                            self._reset_transacao()
                            continue

                except KeyError as e_key:
                    # Coluna inexistente no DataFrame (ex: "CARRO", "OP", "OPERAÇÃO")
                    print(f"❌ Coluna não encontrada no DataFrame: {e_key}")
                    continue
                except Exception as e_linha:
                    # Erro genérico inesperado: registra e continua para a próxima linha
                    print(f"❌ Erro inesperado ao processar linha {index + 1}: {e_linha}")
                    status_msg = "ORDEM NÃO SINCRONIZADA NO MES, ACIONAR PPCP"
                    self._atualizar_status(data, index, arquivo_excel, status_msg)
                    self._reset_transacao()
                    continue

            print("\n" + "=" * 80)
            print("PROCESSAMENTO DE ORDENS FINALIZADO")
            print("=" * 80)

            # Executa fechamento final somente após processar todas as linhas
            try:
                print("\nExecutando fechamento final: ZLOBMM001...")
                self.ZLOBMM001()
                print("✓ ZLOBMM001 executada com sucesso")
            except Exception as e_zlob:
                # Não interrompe o resumo final do robô caso a etapa de fechamento falhe
                print(f"❌ Falha ao executar ZLOBMM001: {e_zlob}")

            # Última tentativa para garantir persistência de status com falha transitória.
            self._persistir_status_pendentes(data, arquivo_excel)

            # -----------------------------------------------------------------
            # CONTAGEM FINAL — relê o Excel salvo para garantir dados atuais
            # "SOLICITADA" cobre: "ORDEM SOLICITADA!" e "ORDEM SOLICITADA VIA ZLOLMM027"
            # -----------------------------------------------------------------
            data_final = pd.read_excel(arquivo_excel, sheet_name="DataBase").astype(str)
            total   = len(data_final)
            sucesso = int(data_final["STATUS"].str.contains("SOLICITADA", na=False).sum())
            erros   = total - sucesso

            print(f"\n  Total processadas : {total}")
            print(f"  Com sucesso       : {sucesso}")
            print(f"  Com erro          : {erros}")

            return total, sucesso, erros

        except Exception as e_geral:
            messagebox.showerror("Erro", f"Erro ao processar planilha: {e_geral}")
            print(f"❌ Erro geral: {e_geral}")
            return 0, 0, 0

    # =========================================================================
    # LOGIN NO SAP
    # =========================================================================

    def sapLogin(self):
        """
        Conclui o fluxo de login interativo do SAP e inicia o processamento.

        Etapas:
            1. Aguarda a tela de login carregar
            2. Confirma os popups iniciais de autenticação
            3. Delega o processamento das ordens a processar_ordem_com_fluxo()

        Retorna:
            tuple(int, int, int): (total, sucesso, erros) propagado do orquestrador.
        """
        try:
            time.sleep(0.5)
            self.session.findById("wnd[0]").sendVKey(0)               # Confirma tela inicial
            self.session.findById("wnd[1]/usr/btnENTER").press()      # Botão Enter no popup de login
            self.session.findById("wnd[1]/usr/btnSEL_BUTTON").press() # Seleciona o perfil de usuário
        except:
            print(f"Erro no login: {sys.exc_info()[0]}")

        time.sleep(1)  # Aguarda o SAP finalizar o processo de login

        # Delega todo o processamento ao orquestrador e retorna as contagens para a UI
        return self.processar_ordem_com_fluxo()


    # =========================================================================
    # TRANSAÇÃO ZDPQPL126 — MONTADOR
    # =========================================================================

    def ZDPQPL126_MONTADOR(self, row, index, data):
        """
        Processa uma única linha de ordem do tipo MONTADOR.

        Etapas:
            1. Acessa ZDPQPL126 e preenche com os dados da OP
            2. Filtra pelo número de operação (VORNR_OP) formatado com zfill(4)
            3. Exporta o resultado para C:\\base\\ZDPQPL126.txt
            4. Lê o TXT e extrai o valor da Linha 7, Coluna 5 (data de início)
            5. Captura o valor adicional da shell (grid) do SAP
            6. Acessa ZLOLMM025 com linha "M" e preenche os campos do Montador
            7. Confirma os popups de finalização da solicitação
            8. Grava STATUS = "ORDEM SOLICITADA!" no Excel

        Parâmetros:
            row   (Series)    : Linha atual do DataFrame com CARRO, OP, OPERAÇÃO, STATUS.
            index (int)       : Índice da linha no DataFrame.
            data  (DataFrame) : DataFrame completo usado para atualizar o STATUS.

        Em caso de falha em qualquer etapa crítica:
            - Grava o STATUS de erro correspondente no Excel
            - Reseta o SAP com _reset_transacao()
            - Encerra via 'return' (sem relançar, pois o orquestrador trata o erro)
        """
        arquivo_excel = self.EXCEL_PATH
        data_126 = []  # Lista para armazenar: [valor_do_TXT, valor_da_shell]

        print(index, row["CARRO"], row["OP"], row["OPERAÇÃO"], row["STATUS"])

        # Formata o número da operação com zeros à esquerda até 4 dígitos
        # Exemplo: "200" → "0200" | "10" → "0010"
        op_val = str(row["OPERAÇÃO"]).strip().zfill(4)
        print(f"Valor formatado da operação: '{op_val}' (len={len(op_val)})")

        # ------------------------------------------------------------------
        # PASSO 1 — Acessar ZDPQPL126 e informar a Ordem de Produção (OP)
        # ------------------------------------------------------------------
        try:
            self.session.findById("wnd[0]/tbar[0]/okcd").text = "/NZDPQPL126"
            self.session.findById("wnd[0]").sendVKey(0)
            self.session.findById("wnd[0]/usr/txtSP$00003-LOW").text = row["OP"]
            self.session.findById("wnd[0]/usr/ctxt%LAYOUT").text = "/FRBARRO"
            self.session.findById("wnd[0]/usr/ctxt%LAYOUT").setFocus()
            self.session.findById("wnd[0]/usr/ctxt%LAYOUT").caretPosition = 8
            self.session.findById("wnd[0]").sendVKey(0)
        except:
            status_msg = "ORDEM NÃO SINCRONIZADA NO MES, ACIONAR PPCP"
            log_msg = f"Falha ao acessar transação ZDPQPL126: {status_msg}"
            self._tratar_falha_e_continuar(data, index, arquivo_excel, data_126, status_msg, log_msg)
            return  # Encerra o processamento desta linha sem relançar

        # ------------------------------------------------------------------
        # PASSO 2 — Filtrar pela operação e exportar o resultado para TXT
        # ------------------------------------------------------------------
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").setCurrentCell(-1, "VORNR_OP")
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectColumn("VORNR_OP")
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarButton("&MB_FILTER")
        self.session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").text = op_val
        self.session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").caretPosition = 4
        self.session.findById("wnd[1]/tbar[0]/btn[0]").press()
        # Exporta o conteúdo filtrado para arquivo TXT local
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarContextButton("&MB_EXPORT")
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectContextMenuItem("&PC")
        self.session.findById("wnd[1]/tbar[0]/btn[0]").press()
        self.session.findById("wnd[1]/usr/ctxtDY_PATH").text = "C:\\base"
        self.session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ZDPQPL126.txt"
        self.session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 13
        self.session.findById("wnd[1]/tbar[0]/btn[11]").press()

        # ------------------------------------------------------------------
        # PASSO 3 — Ler o TXT exportado e extrair Linha 7, Coluna 5
        #           (contém a data/período de início da operação)
        # ------------------------------------------------------------------
        time.sleep(2)  # Aguarda o SAP finalizar a gravação do arquivo no disco
        caminho_arquivo = self.TXT_EXPORT_PATH

        try:
            if os.path.exists(caminho_arquivo):
                dados_arquivo = []
                with open(caminho_arquivo, 'r', encoding='latin-1') as arquivo_txt:
                    leitor = csv.reader(arquivo_txt, delimiter="|")
                    for linha in leitor:
                        dados_arquivo.append(linha)

                # Linha 7 = índice 6 (zero-based); Coluna 5 = índice 4 (zero-based)
                if len(dados_arquivo) >= 7:
                    linha_7 = dados_arquivo[6]
                    if len(linha_7) >= 5:
                        valor_extraido = linha_7[4].strip()
                        print(f"\n✓ Valor extraído — Linha 7, Coluna 5: {valor_extraido}")
                        data_126.append(valor_extraido)
                    else:
                        # Coluna 5 ausente indica operação não configurada corretamente no MES
                        status_msg = "OPERAÇÃO INCORRETA P2S, ACIONAR PPCP"
                        print(f"❌ Coluna 5 não existe na Linha 7")
                        self._atualizar_status(data, index, arquivo_excel, status_msg)
                        self._reset_transacao()
                        return
                else:
                    print(f"❌ Linha 7 não existe no arquivo (total de linhas: {len(dados_arquivo)})")
            else:
                print(f"❌ Arquivo TXT não encontrado: {caminho_arquivo}")
        except Exception as e:
            print(f"❌ Erro ao ler o arquivo TXT: {e}")

        # ------------------------------------------------------------------
        # PASSO 4 — Capturar o valor adicional exibido na shell (grid) do SAP
        # ------------------------------------------------------------------
        valor = self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").text
        print(f"✓ Valor capturado da shell SAP: {valor}")
        data_126.append(valor)

        # ------------------------------------------------------------------
        # PASSO 5 — Navegar para ZLOLMM025 e configurar os parâmetros
        #           da linha M (Montador): Centro, AS e Linha de produção
        # ------------------------------------------------------------------
        self.session.findById("wnd[0]/tbar[0]/okcd").text = "/NZLOLMM025"
        self.session.findById("wnd[0]").sendVKey(0)

        try:
            self.session.findById("wnd[0]/usr/ctxtS_CENTRO-LOW").text = "BOT1"
            self.session.findById("wnd[0]/usr/ctxtP_AS").text = "AS-B28"
            self.session.findById("wnd[0]/usr/ctxtP_LINHA").text = "M"  # M = Montador
            self.session.findById("wnd[0]/usr/ctxtP_LINHA").setFocus()
            self.session.findById("wnd[0]/usr/ctxtP_LINHA").caretPosition = 1
            self.session.findById("wnd[0]").sendVKey(0)
        except Exception as e:
            status_msg = "ORDEM CONGELADA, ACIONAR PPCP"
            log_msg = f"Falha ao preencher transacao ZLOLMM025: {status_msg}"
            self._tratar_falha_e_continuar(data, index, arquivo_excel, data_126, status_msg, log_msg)
            return

        # ------------------------------------------------------------------
        # PASSO 6 — Preencher os campos detalhados da ordem e executar
        # ------------------------------------------------------------------
        try:
            time.sleep(1)
            self.session.findById("wnd[0]/usr/ctxtS_PROGR2-LOW").text = "0200"
            self.session.findById("wnd[0]/usr/txtP_TAKT2").text = "1"
            self.session.findById("wnd[0]/usr/ctxtP_INICA2").text = data_126[0]  # Data extraída do TXT
            self.session.findById("wnd[0]/usr/ctxtS_ORDEM2-LOW").text = row["OP"]
            self.session.findById("wnd[0]/usr/ctxtS_ORDEM2-LOW").setFocus()
            self.session.findById("wnd[0]/usr/ctxtS_ORDEM2-LOW").caretPosition = 8
            self.session.findById("wnd[0]").sendVKey(0)
            self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
            # Pressiona o botão de seleção na tabela 3 vezes (comportamento esperado pelo SAP)
            self.session.findById("wnd[0]/usr/tblZLOLMM025TC_TL100/btnT_TL100-D01[5,0]").setFocus()
            self.session.findById("wnd[0]/usr/tblZLOLMM025TC_TL100/btnT_TL100-D01[5,0]").press()
            self.session.findById("wnd[0]/usr/tblZLOLMM025TC_TL100/btnT_TL100-D01[5,0]").press()
            self.session.findById("wnd[0]/usr/tblZLOLMM025TC_TL100/btnT_TL100-D01[5,0]").press()
            self.session.findById("wnd[0]/tbar[1]/btn[25]").press()
        except Exception as e:
            status_msg = "ORDEM CONGELADA, ACIONAR PPCP"
            log_msg = f"Falha ao preencher transacao ZLOLMM025: {status_msg}"
            self._tratar_falha_e_continuar(data, index, arquivo_excel, data_126, status_msg, log_msg)
            return

        # ------------------------------------------------------------------
        # PASSO 7 — Confirmar os popups de finalização (até 10 tentativas)
        # ------------------------------------------------------------------
        confirmacoes = 0
        for tentativa in range(10):
            try:
                self.session.findById("wnd[1]/tbar[0]/btn[0]").press()
                confirmacoes += 1
                print(f"Confirmação no popup executada ({tentativa + 1}/10)")
                time.sleep(0.3)
            except Exception:
                break  # Nenhum popup adicional disponível

        if confirmacoes == 0:
            # Nenhum popup foi confirmado: a solicitação não foi enviada
            print("Falha ao confirmar popup.")
            self._reset_transacao()
            return

        # Grava o STATUS de sucesso e finaliza a linha
        status_msg = "ORDEM SOLICITADA!"
        self._atualizar_status(data, index, arquivo_excel, status_msg)
        print(f"STATUS atualizado: {status_msg}")
        self._reset_transacao()
        print("✓ Linha processada com sucesso (MONTADOR)")



    # =========================================================================
    # TRANSAÇÃO ZDPQPL126 — FABRICANTE
    # =========================================================================

    def ZDPQPL126_FABRICANTE(self, row, index, data):
        """
        Processa uma única linha de ordem do tipo FABRICANTE.

        O fluxo é análogo ao ZDPQPL126_MONTADOR, com duas diferenças:
            • P_LINHA = "F" (Fabricante) em vez de "M" (Montador)
            • Campos de detalhe usam IDs específicos da aba Fabricante:
              txtP_TAKT4, ctxtS_PERID4-LOW, ctxtS_ORDEM4-LOW

        Diferença crítica de comportamento em relação ao fluxo anterior:
            Em caso de falha, esta função atualiza o STATUS no Excel e
            retorna False, sem relançar exceção para fallback.

        Parâmetros:
            row   (Series)    : Linha atual do DataFrame com CARRO, OP, OPERAÇÃO, STATUS.
            index (int)       : Índice da linha no DataFrame.
            data  (DataFrame) : DataFrame completo usado para atualizar o STATUS.

        Retorna:
            bool: True quando a ordem é solicitada com sucesso, False em falha.
        """
        arquivo_excel = self.EXCEL_PATH
        data_126 = []  # Lista para armazenar: [valor_do_TXT, valor_da_shell]

        print(index, row["CARRO"], row["OP"], row["OPERAÇÃO"], row["STATUS"])

        # Formata o número da operação com zeros à esquerda até 4 dígitos
        op_val = str(row["OPERAÇÃO"]).strip().zfill(4)
        print(f"Valor formatado da operação: '{op_val}' (len={len(op_val)})")

        # ------------------------------------------------------------------
        # PASSO 1 — Acessar ZDPQPL126 e informar a Ordem de Produção (OP)
        # ------------------------------------------------------------------
        try:
            time.sleep(1)  # Pequena pausa para garantir que o SAP esteja pronto para o próximo comando
            self.session.findById("wnd[0]/tbar[0]/okcd").text = "/NZDPQPL126"
            self.session.findById("wnd[0]").sendVKey(0)
            self.session.findById("wnd[0]/usr/txtSP$00003-LOW").text = row["OP"]
            self.session.findById("wnd[0]/usr/ctxt%LAYOUT").text = "/FRBARRO"
            self.session.findById("wnd[0]/usr/ctxt%LAYOUT").setFocus()
            self.session.findById("wnd[0]/usr/ctxt%LAYOUT").caretPosition = 8
            self.session.findById("wnd[0]").sendVKey(0)
        except:
            time.sleep(1) 
            status_msg = "ORDEM NÃO SINCRONIZADA NO MES, ACIONAR PPCP"
            log_msg = f"Falha ao acessar transação ZDPQPL126: {status_msg}"
            self._tratar_falha_e_continuar(data, index, arquivo_excel, data_126, status_msg, log_msg)
            return False

        # ------------------------------------------------------------------
        # PASSO 2 — Filtrar pela operação e exportar o resultado para TXT
        # ------------------------------------------------------------------
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").setCurrentCell(-1, "VORNR_OP")
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectColumn("VORNR_OP")
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarButton("&MB_FILTER")
        self.session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").text = op_val
        self.session.findById("wnd[1]/usr/ssub%_SUBSCREEN_FREESEL:SAPLSSEL:1105/ctxt%%DYN001-LOW").caretPosition = 4
        self.session.findById("wnd[1]/tbar[0]/btn[0]").press()
        # Exporta o conteúdo filtrado para arquivo TXT local
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").pressToolbarContextButton("&MB_EXPORT")
        self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").selectContextMenuItem("&PC")
        self.session.findById("wnd[1]/tbar[0]/btn[0]").press()
        self.session.findById("wnd[1]/usr/ctxtDY_PATH").text = "C:\\base"
        self.session.findById("wnd[1]/usr/ctxtDY_FILENAME").text = "ZDPQPL126.txt"
        self.session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 13
        self.session.findById("wnd[1]/tbar[0]/btn[11]").press()

        # ------------------------------------------------------------------
        # PASSO 3 — Ler o TXT exportado e extrair Linha 7, Coluna 5
        # ------------------------------------------------------------------
        time.sleep(2)  # Aguarda o SAP finalizar a gravação do arquivo no disco
        caminho_arquivo = self.TXT_EXPORT_PATH

        try:
            if os.path.exists(caminho_arquivo):
                dados_arquivo = []
                with open(caminho_arquivo, 'r', encoding='latin-1') as arquivo_txt:
                    leitor = csv.reader(arquivo_txt, delimiter="|")
                    for linha in leitor:
                        dados_arquivo.append(linha)

                # Linha 7 = índice 6 (zero-based); Coluna 5 = índice 4 (zero-based)
                if len(dados_arquivo) >= 7:
                    linha_7 = dados_arquivo[6]
                    if len(linha_7) >= 5:
                        valor_extraido = linha_7[4].strip()
                        print(f"\n✓ Valor extraído — Linha 7, Coluna 5: {valor_extraido}")
                        data_126.append(valor_extraido)
                    else:
                        # Coluna 5 ausente: operação não configurada corretamente no MES
                        status_msg = "OPERAÇÃO INCORRETA P2S, ACIONAR PPCP"
                        print(f"❌ Coluna 5 não existe na Linha 7")
                        self._atualizar_status(data, index, arquivo_excel, status_msg)
                        self._reset_transacao()
                        return False
                else:
                    status_msg = "ARQUIVO TXT INCOMPLETO, ACIONAR PPCP"
                    print(f"❌ Linha 7 não existe no arquivo (total de linhas: {len(dados_arquivo)})")
                    self._atualizar_status(data, index, arquivo_excel, status_msg)
                    self._reset_transacao()
                    return False
            else:
                status_msg = "ARQUIVO TXT NÃO ENCONTRADO, ACIONAR PPCP"
                print(f"❌ Arquivo TXT não encontrado: {caminho_arquivo}")
                self._atualizar_status(data, index, arquivo_excel, status_msg)
                self._reset_transacao()
                return False
        except Exception as e:
            status_msg = "ERRO NA LEITURA DO TXT, ACIONAR PPCP"
            print(f"❌ Erro ao ler o arquivo TXT: {e}")
            self._atualizar_status(data, index, arquivo_excel, status_msg)
            self._reset_transacao()
            return False

        # ------------------------------------------------------------------
        # PASSO 4 — Capturar o valor adicional exibido na shell (grid) do SAP
        # ------------------------------------------------------------------
        valor = self.session.findById("wnd[0]/usr/cntlCONTAINER/shellcont/shell").text
        print(f"✓ Valor capturado da shell SAP: {valor}")
        data_126.append(valor)

        # ------------------------------------------------------------------
        # PASSO 5 — Navegar para ZLOLMM025 e configurar os parâmetros
        #           da linha F (Fabricante): Centro, AS e Linha de produção
        # ------------------------------------------------------------------
        self.session.findById("wnd[0]/tbar[0]/okcd").text = "/NZLOLMM025"
        self.session.findById("wnd[0]").sendVKey(0)

        try:
            self.session.findById("wnd[0]/usr/ctxtS_CENTRO-LOW").text = "BOT1"
            self.session.findById("wnd[0]/usr/ctxtP_AS").text = "AS-B28"
            self.session.findById("wnd[0]/usr/ctxtP_LINHA").text = "F"  # F = Fabricante
            self.session.findById("wnd[0]/usr/ctxtP_LINHA").setFocus()
            self.session.findById("wnd[0]/usr/ctxtP_LINHA").caretPosition = 1
            self.session.findById("wnd[0]").sendVKey(0)
        except Exception as e:
            status_msg = "ORDEM CONGELADA, ACIONAR PPCP"
            log_msg = f"Falha ao preencher transacao ZLOLMM025: {status_msg}"
            self._tratar_falha_e_continuar(data, index, arquivo_excel, data_126, status_msg, log_msg)
            return False

        # ------------------------------------------------------------------
        # PASSO 6 — Preencher os campos detalhados da ordem e executar
        #           (IDs específicos para a aba Fabricante: TAKT4, PERID4, ORDEM4)
        # ------------------------------------------------------------------
        try:
            time.sleep(1)
            self.session.findById("wnd[0]/usr/txtP_TAKT4").text = "1"
            self.session.findById("wnd[0]/usr/ctxtS_PERID4-LOW").text = data_126[0]  # Data extraída do TXT
            self.session.findById("wnd[0]/usr/ctxtS_ORDEM4-LOW").text = row["OP"]
            self.session.findById("wnd[0]/usr/ctxtS_ORDEM4-LOW").setFocus()
            self.session.findById("wnd[0]/usr/ctxtS_ORDEM4-LOW").caretPosition = 8
            self.session.findById("wnd[0]").sendVKey(0)
            self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
            # Pressiona o botão de seleção na tabela 3 vezes (comportamento esperado pelo SAP)
            self.session.findById("wnd[0]/usr/tblZLOLMM025TC_TL100/btnT_TL100-D01[5,0]").setFocus()
            self.session.findById("wnd[0]/usr/tblZLOLMM025TC_TL100/btnT_TL100-D01[5,0]").press()
            self.session.findById("wnd[0]/usr/tblZLOLMM025TC_TL100/btnT_TL100-D01[5,0]").press()
            self.session.findById("wnd[0]/usr/tblZLOLMM025TC_TL100/btnT_TL100-D01[5,0]").press()
            self.session.findById("wnd[0]/tbar[1]/btn[25]").press()
        except Exception as e:
            status_msg = "ORDEM CONGELADA, ACIONAR PPCP"
            log_msg = f"Falha ao preencher transacao ZLOLMM025: {status_msg}"
            self._tratar_falha_e_continuar(data, index, arquivo_excel, data_126, status_msg, log_msg)
            return False

        # ------------------------------------------------------------------
        # PASSO 7 — Confirmar os popups de finalização (até 10 tentativas)
        # ------------------------------------------------------------------
        confirmacoes = 0
        for tentativa in range(10):
            try:
                self.session.findById("wnd[1]/tbar[0]/btn[0]").press()
                confirmacoes += 1
                print(f"Confirmação no popup executada ({tentativa + 1}/10)")
                time.sleep(0.3)
            except Exception:
                break  # Nenhum popup adicional disponível

        if confirmacoes == 0:
            # Nenhum popup confirmado: a solicitação não foi enviada
            status_msg = "FALHA AO CONFIRMAR POPUP, ACIONAR PPCP"
            print("Falha ao confirmar popup.")
            self._atualizar_status(data, index, arquivo_excel, status_msg)
            self._reset_transacao()
            return False

        # Grava o STATUS de sucesso e finaliza a linha
        status_msg = "ORDEM SOLICITADA!"
        self._atualizar_status(data, index, arquivo_excel, status_msg)
        print(f"STATUS atualizado: {status_msg}")
        self._reset_transacao()
        print("✓ Linha processada com sucesso (FABRICANTE)")
        return True



    # =========================================================================
    # TRANSAÇÃO ZLOLMM027 — FALLBACK MTS
    # =========================================================================

    def ZLOLMM027_MTS(self, row, index, data):
        """
        Processa uma única linha via transação ZLOLMM027 (fallback MTS).

        Este método é chamado exclusivamente como fallback por processar_ordem_com_fluxo()
        quando ZDPQPL126_FABRICANTE falha. Preenche os campos da transação ZLOLMM027
        (código do carro, número da OP e operação) e confirma o popup de solicitação.

        Comportamento em caso de falha:
            1. Atualiza o STATUS com mensagem de falha MTS
            2. Executa um reset manual adicional (/N) para garantir estado limpo do SAP
            3. Relança a exceção para que processar_ordem_com_fluxo() mantenha
               o status de falha de ZDPQPL126_FABRICANTE (não sobrescreve com MTS)

        Parâmetros:
            row   (Series)    : Linha atual do DataFrame com CARRO, OP, OPERAÇÃO, STATUS.
            index (int)       : Índice da linha no DataFrame.
            data  (DataFrame) : DataFrame completo usado para atualizar o STATUS.

        Lança:
            Exception: Qualquer falha é relançada para o orquestrador.
        """
        arquivo_excel = self.EXCEL_PATH
        data_126 = []  # Mantido por consistência com os outros métodos (não utilizado aqui)

        print(index, row["CARRO"], row["OP"], row["OPERAÇÃO"], row["STATUS"])

        # Formata a operação com zeros à esquerda até 4 dígitos
        op_val = str(row["OPERAÇÃO"]).strip().zfill(4)
        print(f"Valor formatado da operação: '{op_val}' (len={len(op_val)})")

        # ------------------------------------------------------------------
        # Preencher os campos e executar a transação ZLOLMM027
        # Campos: P_PICK (código do carro), P_AUFNR (OP), S_VORNR-LOW (operação)
        # ------------------------------------------------------------------
        try:
            self.session.findById("wnd[0]/tbar[0]/okcd").text = "/NZLOLMM027"
            self.session.findById("wnd[0]").sendVKey(0)
            self.session.findById("wnd[0]/usr/ctxtP_PICK").text = row["CARRO"]
            self.session.findById("wnd[0]/usr/ctxtP_AUFNR").text = row["OP"]
            self.session.findById("wnd[0]/usr/txtS_VORNR-LOW").text = row["OPERAÇÃO"]
            self.session.findById("wnd[0]/usr/txtS_VORNR-LOW").setFocus()
            self.session.findById("wnd[0]/usr/txtS_VORNR-LOW").caretPosition = 4
            self.session.findById("wnd[0]").sendVKey(0)
            self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
            self.session.findById("wnd[1]/tbar[0]/btn[0]").press()  # Confirma o popup de solicitação
        except:
            status_msg = "FALHA NA SOLICITAÇÃO ORDEM ZTMS, ACIONAR PPCP"
            log_msg = f"Falha ao acessar transação ZLOLMM027: {status_msg}"
            # Trata a falha: atualiza STATUS, limpa data_126 e reseta a transação
            self._tratar_falha_e_continuar(data, index, arquivo_excel, data_126, status_msg, log_msg)
            # Reset adicional explícito: garante retorno à tela inicial mesmo se
            # _tratar_falha_e_continuar não conseguiu completar o reset normalmente
            self.session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
            self.session.findById("wnd[0]").sendVKey(0)
            raise  # Relança para o orquestrador manter o status de falha do FABRICANTE

        # Nota: o STATUS de sucesso ("ORDEM SOLICITADA VIA ZLOLMM027") NÃO é gravado aqui.
        # É responsabilidade de processar_ordem_com_fluxo() gravá-lo após o retorno bem-sucedido.
        print("✓ Linha processada com sucesso (ZLOLMM027 / MTS)")

# =============================================================================
# INTERFACE GRÁFICA (TKINTER) — Ponto de entrada do programa
# =============================================================================

    def ZLOBMM001(self): 

        self.session.findById("wnd[0]/tbar[0]/okcd").text = "/NZLOBMM001"
        self.session.findById("wnd[0]").sendVKey(0)
        self.session.findById("wnd[0]/tbar[1]/btn[17]").press()
        self.session.findById("wnd[1]/usr/cntlALV_CONTAINER_1/shellcont/shell").selectedRows = "0"
        self.session.findById("wnd[1]/usr/cntlALV_CONTAINER_1/shellcont/shell").doubleClickCurrentCell()
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
        self.session.findById("wnd[0]/tbar[0]/okcd").text = "/N"
        self.session.findById("wnd[0]").sendVKey(0)

if __name__ == '__main__':
    # -------------------------------------------------------------------------
    # Janela principal com tamanho fixo expandido:
    #   • Label de status indica o estado atual (aguardando / em execução / concluído)
    #   • 3 labels de resumo (total / sucesso / erro) são limpas a cada nova execução
    #   • Botão é desabilitado durante o processamento e reabilitado ao terminar,
    #     permitindo iniciar uma nova demanda sem precisar reabrir o programa.
    # -------------------------------------------------------------------------
    window = Tk()
    window.title("Robô SAP - ZLOLMM025")
    window.geometry("430x195")
    window.resizable(False, False)

    # --- Label de status dinâmico ---
    lbl_status = Label(
        window,
        text="Pressione 'Iniciar' para processar a planilha.",
        font=("Segoe UI", 10),
        fg="gray"
    )
    lbl_status.pack(pady=(14, 2))

    # --- Labels de resumo (inicialmente vazias) ---
    lbl_total   = Label(window, text="", font=("Segoe UI", 11))
    lbl_sucesso = Label(window, text="", font=("Segoe UI", 11), fg="green")
    lbl_erro    = Label(window, text="", font=("Segoe UI", 11), fg="red")
    lbl_total.pack(pady=2)
    lbl_sucesso.pack(pady=2)
    lbl_erro.pack(pady=2)

    def iniciar():
        """
        Callback do botão 'Iniciar / Iniciar nova demanda'.

        Fluxo:
            1. Desabilita o botão e exibe estado "em execução"
            2. Limpa os resultados da execução anterior
            3. Força a atualização visual da janela antes da chamada bloqueante
            4. Instancia SapGui e dispara sapLogin() (que processa todas as linhas)
            5. Ao retornar, exibe o resumo e reabilita o botão para nova demanda
            6. Agenda a abertura do Excel 3 s após o término
        """
        # --- Prepara UI para a execução ---
        botao.config(state="disabled", text="Processando...")
        lbl_status.config(text="Robô em execução, aguarde...", fg="blue")
        lbl_total.config(text="")
        lbl_sucesso.config(text="")
        lbl_erro.config(text="")
        window.update()  # Garante que as alterações visuais apareçam antes do bloqueio

        # --- Executa o processamento ---
        try:
            resultado = SapGui().sapLogin()
            total, sucesso, erros = resultado if resultado else (0, 0, 0)
        except Exception as e:
            messagebox.showerror("Erro", str(e))
            total = sucesso = erros = 0

        # --- Exibe o resumo ---
        lbl_status.config(text="✔ Processamento concluído!", fg="black")
        lbl_total.config(text=f"Total de ordens processadas :  {total}")
        lbl_sucesso.config(text=f"Total de ordens com sucesso :  {sucesso}")
        lbl_erro.config(text=f"Total de ordens com erro        :  {erros}")

        # --- Reabilita o botão para a próxima demanda ---
        botao.config(state="normal", text="Iniciar nova demanda")

        # --- Abre o Excel atualizado após 3 s sem travar a UI ---
        window.after(3000, lambda: os.startfile(r"C:\base\ZLOLMM025.xlsx"))

    botao = Button(window, text="Iniciar", command=iniciar, width=22, height=2)
    botao.pack(pady=(6, 12))
    mainloop()