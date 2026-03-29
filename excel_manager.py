# =============================================================================
# excel_manager.py — Leitura e escrita da planilha de controle (ZLOLMM025.xlsx)
# =============================================================================

import os
import time
from typing import Any

import pandas as pd
import win32com.client
from openpyxl import load_workbook

from config import EXCEL_PATH


class ExcelManager:
    """
    Gerencia toda a I/O da planilha de controle.

    Responsabilidades:
        - Ler o DataFrame da aba 'DataBase'
        - Persistir a coluna STATUS linha a linha via openpyxl (cell-level write)
        - Usar automação COM do Excel como fallback quando o arquivo está bloqueado
        - Manter um buffer de STATUS pendentes para re-tentativa no encerramento

    Parâmetros:
        arquivo_excel (str): Caminho completo do arquivo .xlsx. Padrão: EXCEL_PATH.
    """

    def __init__(self, arquivo_excel: str = EXCEL_PATH) -> None:
        self.arquivo_excel = arquivo_excel
        self._status_pendentes: dict[int, str] = {}

    # -------------------------------------------------------------------------
    # Leitura
    # -------------------------------------------------------------------------

    def ler_planilha(self) -> pd.DataFrame:
        """Lê a aba 'DataBase' e retorna um DataFrame com todas as colunas como str."""
        return pd.read_excel(self.arquivo_excel, sheet_name="DataBase").astype(str)

    # -------------------------------------------------------------------------
    # Helpers internos de openpyxl
    # -------------------------------------------------------------------------

    def _obter_coluna_status(self, worksheet: Any) -> int:
        """
        Localiza a coluna STATUS na worksheet, criando-a se necessário.

        Retorna:
            int: Índice 1-based da coluna STATUS.
        """
        for col in range(1, worksheet.max_column + 1):
            if worksheet.cell(row=1, column=col).value == "STATUS":
                return col
        nova_col = worksheet.max_column + 1
        worksheet.cell(row=1, column=nova_col).value = "STATUS"
        return nova_col

    # -------------------------------------------------------------------------
    # Persistência principal
    # -------------------------------------------------------------------------

    def salvar(self, data: pd.DataFrame) -> bool:
        """
        Persiste somente a coluna STATUS do DataFrame na aba 'DataBase'.

        Tenta até 5 vezes, aguardando 2 s entre cada tentativa.
        Em caso de PermissionError (arquivo aberto), aciona o fallback COM.

        Retorna:
            bool: True se salvo com sucesso, False após todas as tentativas.
        """
        for tentativa in range(1, 6):
            workbook = None
            try:
                workbook = load_workbook(self.arquivo_excel)
                worksheet = workbook["DataBase"]
                col_status = self._obter_coluna_status(worksheet)

                for index, status_msg in data["STATUS"].items():
                    worksheet.cell(row=index + 2, column=col_status).value = status_msg

                workbook.save(self.arquivo_excel)
                return True

            except PermissionError as e:
                print(
                    f"Falha ao salvar STATUS (tentativa {tentativa}/5): "
                    f"arquivo em uso. Erro: {e}"
                )
                if self._salvar_via_com(data):
                    return True
                time.sleep(2)

            except Exception as e:
                print(f"Falha ao salvar STATUS (tentativa {tentativa}/5): {e}")
                time.sleep(2)

            finally:
                if workbook is not None:
                    workbook.close()

        return False

    def _salvar_via_com(self, data: pd.DataFrame) -> bool:
        """
        Fallback: persiste a coluna STATUS via automação COM do Excel.

        Usado quando o arquivo está bloqueado para escrita por openpyxl.

        Retorna:
            bool: True se salvou com sucesso, False em caso de erro.
        """
        excel = None
        workbook = None
        criou_instancia = False
        wb_ja_aberto = False

        try:
            try:
                excel = win32com.client.GetObject(Class="Excel.Application")
            except Exception:
                excel = win32com.client.Dispatch("Excel.Application")
                criou_instancia = True

            caminho_alvo = os.path.abspath(self.arquivo_excel).lower()
            for wb in excel.Workbooks:
                if os.path.abspath(wb.FullName).lower() == caminho_alvo:
                    workbook = wb
                    wb_ja_aberto = True
                    break

            if workbook is None:
                workbook = excel.Workbooks.Open(self.arquivo_excel)

            worksheet = workbook.Worksheets("DataBase")

            col_status = None
            total_cols = max(worksheet.UsedRange.Columns.Count, 1)
            for col in range(1, total_cols + 1):
                if worksheet.Cells(1, col).Value == "STATUS":
                    col_status = col
                    break

            if col_status is None:
                col_status = total_cols + 1
                worksheet.Cells(1, col_status).Value = "STATUS"

            for index, status_msg in data["STATUS"].items():
                worksheet.Cells(index + 2, col_status).Value = status_msg

            workbook.Save()
            print("✓ STATUS salvo via Excel COM.")
            return True

        except Exception as e:
            print(f"Falha ao salvar STATUS via COM: {e}")
            return False

        finally:
            if workbook is not None and not wb_ja_aberto:
                workbook.Close(SaveChanges=True)
            if excel is not None and criou_instancia:
                excel.Quit()

    # -------------------------------------------------------------------------
    # Atualização de linha + diagnóstico
    # -------------------------------------------------------------------------

    def atualizar_status(self, data: pd.DataFrame, index: int, status_msg: str) -> bool:
        """
        Atualiza em memória e persiste o STATUS de uma única linha.

        Registra o status em `_status_pendentes` se a persistência falhar.

        Retorna:
            bool: True se o arquivo foi salvo com sucesso.
        """
        carro  = data.at[index, "CARRO"]  if "CARRO"  in data.columns else "<sem CARRO>"
        op     = data.at[index, "OP"]     if "OP"     in data.columns else "<sem OP>"
        ant    = data.at[index, "STATUS"] if "STATUS" in data.columns else None

        print(
            f"[STATUS] Linha={index + 1} CARRO={carro} OP={op} | "
            f"anterior={ant!r} -> novo={status_msg!r}"
        )

        data.at[index, "STATUS"] = status_msg
        salvo = self.salvar(data)

        if not salvo:
            self._status_pendentes[index] = status_msg
        else:
            self._status_pendentes.pop(index, None)

        status_mem   = data.at[index, "STATUS"] if "STATUS" in data.columns else None
        status_excel = self.ler_status_linha(index)

        print(
            f"[STATUS] Linha={index + 1} resultado={salvo} | "
            f"memoria={status_mem!r} | excel={status_excel!r} | "
            f"pendentes={self._status_pendentes}"
        )
        return salvo

    def ler_status_linha(self, index: int) -> str | None:
        """
        Lê o valor atual de STATUS diretamente do Excel para diagnóstico.

        Retorna:
            str | None: Valor da célula, ou None em caso de erro.
        """
        workbook = None
        try:
            workbook = load_workbook(self.arquivo_excel, data_only=True)
            worksheet = workbook["DataBase"]
            col = self._obter_coluna_status(worksheet)
            return worksheet.cell(row=index + 2, column=col).value
        except Exception as e:
            print(f"⚠ Falha ao reler STATUS para diagnóstico: {e}")
            return None
        finally:
            if workbook is not None:
                workbook.close()

    # -------------------------------------------------------------------------
    # Flush de pendentes
    # -------------------------------------------------------------------------

    def persistir_pendentes(self, data: pd.DataFrame) -> bool:
        """
        Reaplica e persiste STATUS que falharam em tentativas anteriores.

        Retorna:
            bool: True quando todos os pendentes foram gravados.
        """
        if not self._status_pendentes:
            return True

        print(f"↻ Persistindo {len(self._status_pendentes)} STATUS pendente(s)...")
        for index, status_msg in self._status_pendentes.items():
            data.at[index, "STATUS"] = status_msg

        salvo = self.salvar(data)
        if salvo:
            self._status_pendentes.clear()
            print("✓ STATUS pendentes gravados com sucesso.")
            return True

        print("⚠ Ainda há STATUS pendentes não gravados no Excel.")
        return False
